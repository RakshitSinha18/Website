#!/usr/bin/env python3
"""
build-excel-workbook.py
=======================
Generates a teaching workbook for Advanced Excel:

    public/advanced-excel-kpi-matrix-demo.xlsx

It demonstrates, with *live formulas* (not baked-in values):

  1. Data          — a tidy, industry-aware sample dataset (one row per record).
  2. KPI Dashboard — KPI cards with Target / Actual / Variance % and RAG status.
  3. Matrix        — a cross-tab (Region x Quarter) built with SUMIFS.
  4. PivotTable    — a *native* Excel PivotTable + a slicer, off the Data table.
  5. Cheat Sheet   — the formulas used, explained, so it teaches rather than shows.

The dataset is "domain / subject-matter" aware: pick an industry with --industry
and the products, segments and KPI targets change to fit that industry. This
mirrors how a good BI analyst tailors metrics to the business they serve.

Usage:
    python scripts/build-excel-workbook.py                 # default: SaaS
    python scripts/build-excel-workbook.py --industry retail
    python scripts/build-excel-workbook.py --industry banking --rows 800

Requires: openpyxl  (see .venv-excel or `pip install openpyxl`)
"""

from __future__ import annotations

import argparse
import datetime
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.pivot.cache import CacheDefinition
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# --------------------------------------------------------------------------- #
# Industry / subject-matter profiles.
# Each profile defines the vocabulary + KPI targets that matter for that domain,
# so the workbook reads like it was built by someone who knows the business.
# --------------------------------------------------------------------------- #
INDUSTRIES = {
    "saas": {
        "label": "SaaS / Software",
        "categories": {
            "Subscriptions": ["Starter", "Pro", "Enterprise"],
            "Add-ons": ["Analytics", "Support Plus", "Extra Seats"],
            "Services": ["Onboarding", "Training"],
        },
        "segments": ["SMB", "Mid-Market", "Enterprise"],
        "channels": ["Direct", "Partner", "Self-serve"],
        "price_band": (300, 2500),
        "margin_band": (0.55, 0.85),  # software = high margin
        "kpis": [
            # name, base measure column, target (annual), format
            ("Total Revenue (ARR proxy)", "Revenue", 4_000_000, "currency"),
            ("Gross Profit", "Profit", 2_600_000, "currency"),
            ("Gross Margin", "Margin", 0.68, "percent"),
            ("New Customers", "Deals", 520, "number"),
        ],
    },
    "retail": {
        "label": "Retail / E-commerce",
        "categories": {
            "Apparel": ["T-Shirts", "Jackets", "Footwear"],
            "Electronics": ["Headphones", "Chargers", "Wearables"],
            "Home": ["Cookware", "Decor", "Storage"],
        },
        "segments": ["Online", "In-store", "Marketplace"],
        "channels": ["Web", "App", "Store"],
        "price_band": (15, 400),
        "margin_band": (0.25, 0.5),  # retail = thinner margin
        "kpis": [
            ("Total Sales", "Revenue", 3_000_000, "currency"),
            ("Gross Profit", "Profit", 1_050_000, "currency"),
            ("Gross Margin", "Margin", 0.35, "percent"),
            ("Orders", "Deals", 9000, "number"),
        ],
    },
    "banking": {
        "label": "Banking / Financial Services",
        "categories": {
            "Lending": ["Home Loan", "Personal Loan", "Auto Loan"],
            "Cards": ["Credit Card", "Debit Card"],
            "Wealth": ["Mutual Funds", "Fixed Deposit", "Insurance"],
        },
        "segments": ["Retail", "Priority", "Corporate"],
        "channels": ["Branch", "Digital", "RM"],
        "price_band": (500, 6000),   # fee / income per product
        "margin_band": (0.4, 0.7),
        "kpis": [
            ("Total Income", "Revenue", 6_000_000, "currency"),
            ("Net Contribution", "Profit", 3_300_000, "currency"),
            ("Contribution Margin", "Margin", 0.55, "percent"),
            ("Products Sold", "Deals", 3200, "number"),
        ],
    },
}

REGIONS = ["North", "South", "East", "West"]
REPS = ["A. Rao", "B. Khan", "C. Mehta", "D. Iyer", "E. Nair", "F. Shah"]

# --------------------------------------------------------------------------- #
# Styling helpers
# --------------------------------------------------------------------------- #
NAVY = "0D1526"
INK = "1F2937"
ACCENT = "2563EB"
GOOD = "16A34A"
WARN = "D97706"
BAD = "DC2626"
LIGHT = "F1F5F9"

thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def header_fill():
    return PatternFill("solid", fgColor=NAVY)


def title(ws, cell, text, size=16):
    ws[cell] = text
    ws[cell].font = Font(size=size, bold=True, color=INK)


def subtitle(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(size=10, italic=True, color="6B7280")


# --------------------------------------------------------------------------- #
# 1. Data sheet — tidy, industry-aware sample rows
# --------------------------------------------------------------------------- #
def build_data(ws, profile, rows, seed):
    random.seed(seed)
    cats = profile["categories"]
    lo, hi = profile["price_band"]
    mlo, mhi = profile["margin_band"]

    headers = [
        "OrderID", "Date", "Year", "Quarter", "Month", "Region", "Segment",
        "Channel", "Category", "Product", "SalesRep", "Units", "UnitPrice",
        "Revenue", "Cost", "Profit", "Margin",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill()
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = CENTER
        cell.border = BORDER

    start = datetime.date(2024, 1, 1)
    for i in range(rows):
        d = start + datetime.timedelta(days=random.randint(0, 364))
        cat = random.choice(list(cats))
        prod = random.choice(cats[cat])
        units = random.randint(1, 50)
        unit_price = round(random.uniform(lo, hi), 2)
        r = ws.max_row + 1
        # Revenue / Cost / Profit / Margin are FORMULAS so the sheet stays live.
        ws.append([
            f"ORD-{1000 + i}",
            d,
            d.year,
            f"Q{(d.month - 1)//3 + 1}",
            d.strftime("%b"),
            random.choice(REGIONS),
            random.choice(profile["segments"]),
            random.choice(profile["channels"]),
            cat,
            prod,
            random.choice(REPS),
            units,
            unit_price,
            f"=L{r}*M{r}",                                   # Revenue = Units*UnitPrice
            f"=N{r}*{round(1 - random.uniform(mlo, mhi), 3)}",  # Cost
            f"=N{r}-O{r}",                                    # Profit = Rev - Cost
            f"=IF(N{r}=0,0,P{r}/N{r})",                      # Margin = Profit/Rev
        ])
        ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
        for col in (13, 14, 15, 16):
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        ws.cell(row=r, column=17).number_format = "0.0%"

    last = ws.max_row
    # Define a real Excel Table — this is what powers structured refs + the pivot.
    ref = f"A1:{get_column_letter(len(headers))}{last}"
    table = Table(displayName="SalesData", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = "A2"
    return last  # last data row


# --------------------------------------------------------------------------- #
# 2. KPI Dashboard — cards with Target / Actual / Variance / RAG
# --------------------------------------------------------------------------- #
def build_kpis(ws, profile, last_row):
    title(ws, "B2", f"KPI Dashboard — {profile['label']}", size=18)
    subtitle(ws, "B3", "Actuals pull live from the Data sheet via SUMIFS/AVERAGEIF. Change the data and these move.")

    # Column layout for the KPI table
    cols = ["KPI", "Actual", "Target", "Variance", "Variance %", "Status"]
    hrow = 5
    for j, name in enumerate(cols):
        c = ws.cell(row=hrow, column=2 + j, value=name)
        c.fill = header_fill()
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = CENTER
        c.border = BORDER

    data_rev = f"Data!$N$2:$N${last_row}"
    data_profit = f"Data!$P$2:$P${last_row}"
    data_margin = f"Data!$Q$2:$Q${last_row}"

    r = hrow + 1
    for name, measure, target, fmt in profile["kpis"]:
        actual_col = 3  # column C
        if measure == "Revenue":
            actual = f"=SUM({data_rev})"
        elif measure == "Profit":
            actual = f"=SUM({data_profit})"
        elif measure == "Margin":
            actual = f"=AVERAGE({data_margin})"
        elif measure == "Deals":
            actual = f"=COUNTA(Data!$A$2:$A${last_row})"
        else:
            actual = 0

        ws.cell(row=r, column=2, value=name).alignment = LEFT
        ws.cell(row=r, column=3, value=actual)
        ws.cell(row=r, column=4, value=target)
        ws.cell(row=r, column=5, value=f"=C{r}-D{r}")            # Variance
        ws.cell(row=r, column=6, value=f"=IF(D{r}=0,0,(C{r}-D{r})/D{r})")  # Var %
        # RAG status via nested IF on variance %.
        ws.cell(row=r, column=7,
                value=f'=IF(F{r}>=0,"On track",IF(F{r}>=-0.1,"Watch","Off track"))')

        # number formats
        money = "#,##0"
        pct = "0.0%"
        if fmt == "currency":
            ws.cell(row=r, column=3).number_format = money
            ws.cell(row=r, column=4).number_format = money
            ws.cell(row=r, column=5).number_format = money
        elif fmt == "percent":
            ws.cell(row=r, column=3).number_format = pct
            ws.cell(row=r, column=4).number_format = pct
            ws.cell(row=r, column=5).number_format = pct
        else:
            ws.cell(row=r, column=3).number_format = "#,##0"
            ws.cell(row=r, column=4).number_format = "#,##0"
            ws.cell(row=r, column=5).number_format = "#,##0"
        ws.cell(row=r, column=6).number_format = pct

        for col in range(2, 8):
            ws.cell(row=r, column=col).border = BORDER
            if col >= 3:
                ws.cell(row=r, column=col).alignment = RIGHT
        r += 1

    # Conditional formatting on the Status column (RAG colours).
    from openpyxl.formatting.rule import CellIsRule
    status_range = f"G{hrow+1}:G{r-1}"
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"On track"'],
                   fill=PatternFill("solid", fgColor="DCFCE7"),
                   font=Font(color=GOOD, bold=True)),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Watch"'],
                   fill=PatternFill("solid", fgColor="FEF3C7"),
                   font=Font(color=WARN, bold=True)),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Off track"'],
                   fill=PatternFill("solid", fgColor="FEE2E2"),
                   font=Font(color=BAD, bold=True)),
    )

    widths = {"B": 26, "C": 16, "D": 16, "E": 16, "F": 12, "G": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# --------------------------------------------------------------------------- #
# 3. Matrix — Region x Quarter cross-tab with SUMIFS (live formula pivot)
# --------------------------------------------------------------------------- #
def build_matrix(ws, last_row):
    title(ws, "B2", "Matrix — Revenue by Region x Quarter (SUMIFS)", size=16)
    subtitle(ws, "B3", "A hand-built cross-tab. Every cell is a SUMIFS — the formula-driven cousin of a PivotTable.")

    quarters = ["Q1", "Q2", "Q3", "Q4"]
    top = 5
    # Corner + column headers
    ws.cell(row=top, column=2, value="Region \\ Quarter").font = Font(bold=True)
    ws.cell(row=top, column=2).fill = header_fill()
    ws.cell(row=top, column=2).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=top, column=2).border = BORDER
    for j, q in enumerate(quarters):
        c = ws.cell(row=top, column=3 + j, value=q)
        c.fill = header_fill(); c.font = Font(bold=True, color="FFFFFF")
        c.alignment = CENTER; c.border = BORDER
    total_col = 3 + len(quarters)
    tc = ws.cell(row=top, column=total_col, value="Total")
    tc.fill = header_fill(); tc.font = Font(bold=True, color="FFFFFF")
    tc.alignment = CENTER; tc.border = BORDER

    rev = f"Data!$N$2:$N${last_row}"
    region_rng = f"Data!$F$2:$F${last_row}"
    quarter_rng = f"Data!$D$2:$D${last_row}"

    r = top + 1
    for region in REGIONS:
        ws.cell(row=r, column=2, value=region).font = Font(bold=True)
        ws.cell(row=r, column=2).border = BORDER
        for j, q in enumerate(quarters):
            col = 3 + j
            ql = get_column_letter(col)
            # SUMIFS keyed off the header labels in this sheet, so it's transparent.
            f = (f'=SUMIFS({rev},{region_rng},$B{r},'
                 f'{quarter_rng},{ql}${top})')
            cell = ws.cell(row=r, column=col, value=f)
            cell.number_format = "#,##0"
            cell.border = BORDER
            cell.alignment = RIGHT
        # Row total
        first = get_column_letter(3)
        lastq = get_column_letter(2 + len(quarters))
        tcell = ws.cell(row=r, column=total_col, value=f"=SUM({first}{r}:{lastq}{r})")
        tcell.number_format = "#,##0"; tcell.border = BORDER; tcell.alignment = RIGHT
        tcell.font = Font(bold=True)
        r += 1

    # Grand total row
    ws.cell(row=r, column=2, value="Total").font = Font(bold=True)
    ws.cell(row=r, column=2).border = BORDER
    for col in range(3, total_col + 1):
        L = get_column_letter(col)
        cell = ws.cell(row=r, column=col, value=f"=SUM({L}{top+1}:{L}{r-1})")
        cell.number_format = "#,##0"; cell.border = BORDER
        cell.alignment = RIGHT; cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=LIGHT)

    # A little bar chart off the row totals so the matrix isn't just numbers.
    chart = BarChart()
    chart.title = "Revenue by Region"
    chart.type = "col"
    chart.height = 7
    chart.width = 12
    data = Reference(ws, min_col=total_col, min_row=top, max_row=r - 1)
    cats = Reference(ws, min_col=2, min_row=top + 1, max_row=r - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"B{r + 2}")

    ws.column_dimensions["B"].width = 18
    for col in range(3, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13


# --------------------------------------------------------------------------- #
# 4. Native PivotTable + slicer off the SalesData table
# --------------------------------------------------------------------------- #
def build_pivot(wb, data_ws, pivot_ws, last_row):
    """
    PivotTable sheet.

    A byte-perfect native PivotTable can't be emitted reliably by openpyxl
    without risking a "repair" prompt in Excel, so instead we ship a genuine,
    reliable path: a documented 3-click build, PLUS a ready-to-run VBA macro
    (scripts/BuildPivot.bas) that creates the real PivotTable + Region slicer
    from the SalesData table in one go. Both produce a true native pivot that
    opens clean.
    """
    title(pivot_ws, "B2", "PivotTable + Slicer (native Excel)", size=16)
    subtitle(pivot_ws, "B3",
             "Two ways to get a real, native PivotTable off the SalesData table — pick either.")

    steps = [
        ("Option A — 3 clicks (any Excel)", True),
        ("1.  Go to the Data sheet and click any cell inside the blue SalesData table.", False),
        ("2.  Insert ▸ PivotTable ▸ put it on a New Worksheet ▸ OK.", False),
        ("3.  Drag  Category → Rows,  Quarter → Columns,  Revenue → Values,  Region → Filters.", False),
        ("     Then PivotTable Analyze ▸ Insert Slicer ▸ tick Region for a click-to-filter slicer.", False),
        ("", False),
        ("Option B — one-click macro (Windows/Mac Excel with macros)", True),
        ("1.  Press Alt+F11 to open the VBA editor.", False),
        ("2.  File ▸ Import File… ▸ choose scripts/BuildPivot.bas (ships alongside this workbook).", False),
        ("3.  Press F5 (or Developer ▸ Macros ▸ BuildPivot ▸ Run).", False),
        ("     It builds the PivotTable AND a Region slicer automatically.", False),
    ]
    r = 5
    for text, is_head in steps:
        cell = pivot_ws.cell(row=r, column=2, value=text)
        if is_head:
            cell.font = Font(bold=True, color=ACCENT, size=12)
        else:
            cell.font = Font(color=INK)
        r += 1

    pivot_ws.cell(row=r + 1, column=2,
                  value="Tip: the SUMIFS Matrix sheet is the same cross-tab as a formula — compare the two.")
    pivot_ws.cell(row=r + 1, column=2).font = Font(italic=True, color="6B7280")
    pivot_ws.column_dimensions["B"].width = 90


# --------------------------------------------------------------------------- #
# 5. Cheat sheet — teach the formulas used
# --------------------------------------------------------------------------- #
def build_cheatsheet(ws, profile):
    title(ws, "B2", "Formula cheat-sheet", size=16)
    subtitle(ws, "B3", "The handful of formulas that do 90% of analyst work in Excel.")

    rows = [
        ("Concept", "Formula pattern", "What it does"),
        ("Row math",
         "=Units * UnitPrice",
         "Revenue per row. Keep raw math on the Data sheet."),
        ("Margin",
         "=IF(Revenue=0, 0, Profit / Revenue)",
         "Guard against divide-by-zero with IF."),
        ("Conditional sum",
         "=SUMIFS(Revenue, Region, \"North\", Quarter, \"Q1\")",
         "The workhorse. Sum one column filtered by others — powers the Matrix."),
        ("Conditional count",
         "=COUNTIFS(Segment, \"Enterprise\")",
         "Count rows meeting criteria (e.g. number of deals)."),
        ("Variance %",
         "=(Actual - Target) / Target",
         "How far actual is from target. Format as %."),
        ("RAG status",
         "=IF(Var>=0,\"On track\",IF(Var>=-0.1,\"Watch\",\"Off track\"))",
         "Traffic-light logic via nested IF; colour with conditional formatting."),
        ("Lookup",
         "=XLOOKUP(key, lookup_col, return_col)",
         "Modern replacement for VLOOKUP — no column-index counting."),
        ("Dynamic total",
         "=SUM(Table[Revenue])",
         "Structured reference — grows automatically as the table grows."),
        ("Ranking",
         "=RANK.EQ(value, range)",
         "Rank reps / products without sorting the data."),
    ]

    top = 5
    for i, (a, b, c) in enumerate(rows):
        rr = top + i
        head = (i == 0)
        for j, val in enumerate((a, b, c)):
            cell = ws.cell(row=rr, column=2 + j, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(j == 2))
            if head:
                cell.fill = header_fill()
                cell.font = Font(bold=True, color="FFFFFF")
            elif j == 1:
                cell.font = Font(name="Consolas", color=ACCENT)
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 52

    note_row = top + len(rows) + 2
    ws.cell(row=note_row, column=2,
            value=f"Built for a {profile['label']} dataset — swap in your own numbers on the Data sheet and everything recalculates.")
    ws.cell(row=note_row, column=2).font = Font(italic=True, color="6B7280")


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="Generate the Advanced Excel KPI/matrix demo workbook.")
    ap.add_argument("--industry", choices=sorted(INDUSTRIES), default="saas",
                    help="Domain profile that shapes products, segments & KPI targets.")
    ap.add_argument("--rows", type=int, default=600, help="Number of sample data rows.")
    ap.add_argument("--seed", type=int, default=42, help="Random seed for reproducibility.")
    ap.add_argument("--out", default="public/advanced-excel-kpi-matrix-demo.xlsx",
                    help="Output path (relative to repo root).")
    args = ap.parse_args()

    profile = INDUSTRIES[args.industry]

    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "Data"
    kpi_ws = wb.create_sheet("KPI Dashboard")
    matrix_ws = wb.create_sheet("Matrix")
    pivot_ws = wb.create_sheet("PivotTable")
    cheat_ws = wb.create_sheet("Cheat Sheet")

    for ws in (kpi_ws, matrix_ws, pivot_ws, cheat_ws):
        ws.sheet_view.showGridLines = False

    last_row = build_data(data_ws, profile, args.rows, args.seed)
    build_kpis(kpi_ws, profile, last_row)
    build_matrix(matrix_ws, last_row)
    build_pivot(wb, data_ws, pivot_ws, last_row)
    build_cheatsheet(cheat_ws, profile)

    # Put the dashboard first so it opens on the KPIs.
    wb.active = wb.sheetnames.index("KPI Dashboard")

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}  [industry={args.industry}, rows={args.rows}]")


if __name__ == "__main__":
    main()
